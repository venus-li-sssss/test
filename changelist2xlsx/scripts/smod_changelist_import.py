# -*- coding: utf-8 -*-
"""
SMOD 修改点自动导入工作流

结合 changelist2xlsx 转换 + SMOD API，实现：
    1. 读取修改点 txt，生成同名 xlsx
    2. 从版本号提取项目名称（如 QDM559_STM32G0B0_APP_01.001.01.001_V19 → QDM559）
    3. 登录 SMOD，搜索历史项目确定平台（id_plat_ver）
    4. 若版本号已存在，直接导入 xlsx；否则先新建该版本，再导入 xlsx

依赖：
    pip install requests openpyxl
"""

import argparse
import os
import re
import sys
import time
from collections import Counter
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# 复用已有的 SMOD 客户端
from smod_client import SmodClient, SmodApiError

# =====================================================================
# 交互式输入：缺必填参数时弹窗（无 GUI 时回退到命令行 input）
# =====================================================================
def _input_dialog(fields: List[Tuple[str, str, bool, str]]) -> Optional[Dict[str, str]]:
    """弹窗收集多个字段。

    fields: [(label, key, is_password, default_value), ...]
    返回: {key: value} 或 None（用户取消）。
    """
    result: Dict[str, str] = {}
    try:
        import tkinter as tk
        from tkinter import simpledialog, messagebox
    except Exception:  # 无 tkinter 时回退命令行
        print("未检测到图形界面，使用命令行输入：")
        for label, key, is_password, default in fields:
            prompt = f"{label}"
            if default:
                prompt += f"（默认：{default}）"
            prompt += ": "
            if is_password:
                import getpass
                val = getpass.getpass(prompt) or default
            else:
                val = input(prompt) or default
            result[key] = val.strip()
        return result

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    root.update()

    # 创建统一的多字段对话框
    dlg = tk.Toplevel(root)
    dlg.title("请补充 SMOD 导入所需信息")
    dlg.geometry("480x" + str(90 + len(fields) * 38))
    dlg.resizable(False, False)
    dlg.attributes("-topmost", True)
    dlg.protocol("WM_DELETE_WINDOW", lambda: dlg.destroy())

    entries: List[Any] = []
    for i, (label, key, is_password, default) in enumerate(fields):
        tk.Label(dlg, text=label, anchor="w").grid(row=i, column=0, padx=10, pady=6, sticky="w")
        var = tk.StringVar(value=default or "")
        if is_password:
            ent = tk.Entry(dlg, textvariable=var, show="*", width=45)
        else:
            ent = tk.Entry(dlg, textvariable=var, width=45)
        ent.grid(row=i, column=1, padx=10, pady=6)
        entries.append((key, var))

    def on_ok():
        nonlocal result
        result = {k: v.get().strip() for k, v in entries}
        dlg.destroy()

    def on_cancel():
        nonlocal result
        result = None
        dlg.destroy()

    btn_frame = tk.Frame(dlg)
    btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=12)
    tk.Button(btn_frame, text="确定", width=10, command=on_ok).pack(side="left", padx=10)
    tk.Button(btn_frame, text="取消", width=10, command=on_cancel).pack(side="left", padx=10)

    dlg.focus_force()
    if entries:
        entries[0][1].set(entries[0][1].get())
        # 让第一个输入框获得焦点
        first_entry = entries[0][1]
        dlg.after(50, lambda: first_entry.set(first_entry.get()))
    root.wait_window(dlg)
    root.destroy()
    return result


def prompt_missing_args(args) -> argparse.Namespace:
    """根据已提供的参数，弹窗补全缺失项。"""
    fields: List[Tuple[str, str, bool, str]] = []
    if not getattr(args, "txt", None):
        fields.append(("修改点 txt 路径", "txt", False, ""))
    if not getattr(args, "version", None):
        fields.append(("版本号", "version", False, ""))

    has_token = bool(getattr(args, "token", None))
    has_user = bool(getattr(args, "username", None))
    has_pwd = bool(getattr(args, "password", None))
    if not has_token and not (has_user and has_pwd):
        fields.append(("SMOD/SSO 用户名", "username", False, getattr(args, "username", None) or "Venus.Li@ikotek.com"))
        fields.append(("SMOD/SSO 密码", "password", True, ""))
        fields.append(("或浏览器 access_token（可选，填写则跳过密码）", "token", False, getattr(args, "token", None) or ""))

    if not fields:
        return args

    values = _input_dialog(fields)
    if values is None:
        print("用户取消输入，流程结束。")
        sys.exit(0)

    for k, v in values.items():
        setattr(args, k, v or getattr(args, k, None))
    return args


# =====================================================================
# changelist -> xlsx 转换逻辑（与 changelist2xlsx skill 保持一致）
# =====================================================================
HEADERS = [
    "修改点编号", "修改点描述", "修改点分类", "执行人", "测试结果",
    "评审状态", "备注", "Test-Proposal", "Stress-Test", "HW-Test",
]

CATEGORY_MAP = {
    "new requirements": "Newly-added Customer Requirements",
    "bug fix": "Internal Defect Requirements",
    "inner bug fix": "Internal Defect Requirements",
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="969696")
HEADER_FONT = Font(bold=True)
DESC_ALIGN = Alignment(wrap_text=True, vertical="top")


def read_text(path: str) -> str:
    """优先 utf-8，失败回退 gbk/utf-8-sig。"""
    for enc in ("utf-8", "gbk", "utf-8-sig"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, LookupError):
            continue
    with open(path, "rb") as f:
        return f.read().decode("utf-8", errors="ignore")


def derive_category(block: str) -> Optional[str]:
    for line in block.splitlines():
        m = re.match(r"^\s*<([^>]+)>\s*:\s*(.*)$", line)
        if m and m.group(1).strip().lower() == "change type":
            val = m.group(2).strip().lower()
            for key, label in CATEGORY_MAP.items():
                if val == key or val == key.replace(" ", ""):
                    return label
            if "new requirements" in val:
                return CATEGORY_MAP["new requirements"]
            if "bug fix" in val:
                return CATEGORY_MAP["bug fix"]
    return None


def parse_commits(text: str) -> List[Tuple[str, str, Optional[str]]]:
    commits = []
    for chunk in re.split(r"(?m)^commit ", text)[1:]:
        lines = chunk.splitlines()
        if not lines:
            continue
        commit_hash = lines[0].strip()
        block = "\n".join(lines[1:]).rstrip("\r\n")
        commits.append((commit_hash, block, derive_category(block)))
    return commits


def build_workbook(commits):
    wb = Workbook()
    ws = wb.active
    ws.title = "修改点"
    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")
    for ridx, (commit_hash, block, category) in enumerate(commits, start=2):
        ws.cell(row=ridx, column=1, value=commit_hash)
        desc_cell = ws.cell(row=ridx, column=2, value=block)
        desc_cell.alignment = DESC_ALIGN
        ws.cell(row=ridx, column=3, value=category)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 32
    for col in ("D", "E", "F", "G", "H", "I", "J"):
        ws.column_dimensions[col].width = 14
    return wb


def convert_txt_to_xlsx(txt_path: str, xlsx_path: Optional[str] = None) -> str:
    """txt -> 同名 xlsx，返回 xlsx 路径。"""
    if not os.path.isfile(txt_path):
        raise FileNotFoundError(f"找不到修改点 txt: {txt_path}")
    out_path = xlsx_path or os.path.splitext(txt_path)[0] + ".xlsx"
    text = read_text(txt_path)
    commits = parse_commits(text)
    if not commits:
        raise ValueError("未在 txt 中解析到任何 commit，请确认是 git changelist 格式。")
    wb = build_workbook(commits)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    cat_counter = Counter(c[2] or "(空)" for c in commits)
    print(f"[完成] 已解析 {len(commits)} 个 commit")
    for k, v in cat_counter.items():
        print(f"        {k}: {v}")
    print(f"[输出] {out_path}")
    return out_path


# =====================================================================
# 版本号 / 项目名解析
# =====================================================================
def parse_project_name(version_code: str) -> str:
    """从版本号提取项目名称，取第一个下划线前的前缀。"""
    return version_code.split("_")[0]


def infer_platform_from_history(
    client: SmodClient,
    project_name: str,
    platform_code: Optional[str] = None,
) -> Tuple[int, str]:
    """搜索历史项目，若平台唯一则返回 (id_plat_ver, name_plat_ver)，否则让用户选择。

    策略：
        1. 搜索 project_name（如 QDM559）
        2. 统计历史记录里 name_plat_ver 的分布
        3. 唯一 -> 直接返回；多个 -> 列出让用户输入序号选择
    """
    print(f"\n搜索历史项目以推断平台：keyword={project_name}")
    result = client.search_projects(keyword=project_name, page=1, size=100, platform_code=platform_code)
    records = result.get("records", []) if isinstance(result, dict) else []
    if not records:
        raise SmodApiError(f"未搜索到项目 {project_name} 的历史记录，无法自动推断平台。")

    platform_names = [r.get("name_plat_ver") for r in records if r.get("name_plat_ver")]
    counter = Counter(platform_names)
    print(f"找到 {len(records)} 条历史记录，平台分布：")
    for name, cnt in counter.most_common():
        print(f"  - {name}: {cnt} 条")

    if len(counter) == 1:
        chosen_name = platform_names[0]
        print(f"平台唯一，自动选择：{chosen_name}")
    else:
        choices = list(counter.keys())
        print("检测到多个平台，请选择：")
        for idx, name in enumerate(choices, start=1):
            print(f"  {idx}. {name}")
        while True:
            try:
                sel = int(input("请输入序号: ").strip())
                if 1 <= sel <= len(choices):
                    chosen_name = choices[sel - 1]
                    break
            except ValueError:
                pass
            print("输入无效，请重新输入。")

    # 把平台名称映射到 id_plat_ver
    platforms = client.get_hardware_platforms()
    for p in platforms:
        if str(p.get("name", "")).strip().lower() == chosen_name.strip().lower():
            return p.get("id"), chosen_name
    raise SmodApiError(f"在 /api/simpleHardwarePlatforms 中未找到平台 {chosen_name}")


# =====================================================================
# 描述解析 / AI 生成测试用例
# =====================================================================
# 需求标记：描述里出现这些字段，即视为「有需求」
REQUIREMENT_FIELD_RE = re.compile(
    r"<(Change\s*Type|Solution|Change\s*Reason|Test-Proposal|Root\s*Cause|"
    r"Change\s*Id|Affected\s*Function\s*Name)>",
    re.IGNORECASE,
)
# 变更标题行：形如 <QDM559><运输功能/日志功能>: 新增运输模式和日志功能优化
TITLE_LINE_RE = re.compile(
    r"^\s*<(?P<proj>[^>]+)>\s*(?:<(?P<feat>[^>]+)>)?\s*:\s*(?P<desc>\S.*)$"
)
# 普通字段 <Field>: value
FIELD_LINE_RE = re.compile(r"^\s*<([^>]+)>\s*:\s*(.*)$")
# 去掉 Author/Date 头
AUTHOR_DATE_RE = re.compile(r"^\s*(Author|Date)\s*:.*$", re.MULTILINE)


def strip_author_date(description: str) -> str:
    """去掉 'Author: ...\\nDate: ...' 头，返回正文"""
    return AUTHOR_DATE_RE.sub("", description).strip()


def parse_field_values(description: str) -> Dict[str, str]:
    """解析描述中的 <Field>: value 字段"""
    fields = {}
    for line in description.splitlines():
        m = FIELD_LINE_RE.match(line)
        if m:
            fields[m.group(1).strip().lower()] = m.group(2).strip()
    return fields


def has_requirement(description: str) -> bool:
    """判断修改点是否有需求。

    规则：
      - 描述包含 <Change Type>/<Solution>/<Change Reason> 等需求字段 -> 有
      - 描述含变更标题行（<xxx><feature>: 文字）-> 有
      - 其余（典型如只有「版本号 + 0x十六进制」且无任何 <> 标签）-> 无需求
    """
    body = strip_author_date(description)
    if not body:
        return False
    if REQUIREMENT_FIELD_RE.search(body):
        return True
    if TITLE_LINE_RE.match(body) or re.search(r">\s*:\s*\S", body):
        return True
    # 没有任何 <> 标签，且看起来只是版本号/十六进制 -> 无需求
    if "<" not in body and ">" not in body:
        return False
    return False


def extract_title(description: str) -> Optional[Tuple[str, str]]:
    """从描述提取 (features, desc_text)。

    例：<QDM559><运输功能/日志功能>:新增运输模式和日志功能优化
        -> ('运输功能/日志功能', '新增运输模式和日志功能优化')
    若无标题行则返回 ('', <Solution/Change Reason 文本>)
    """
    body = strip_author_date(description)
    # 逐行匹配（避免 $ 锚定整段文本导致多行描述匹配失败）
    for line in body.splitlines():
        m = TITLE_LINE_RE.match(line)
        if m:
            return m.group("feat").strip(), m.group("desc").strip()
    fields = parse_field_values(description)
    for key in ("solution", "change reason", "root cause"):
        if fields.get(key):
            return "", fields[key]
    return None


def _extract_context(description: str) -> Dict[str, Any]:
    """从描述中提取变更标题、功能模块、关键改动上下文。"""
    body = strip_author_date(description)
    fields = parse_field_values(description)

    title_feature, title_desc = "", ""
    for line in body.splitlines():
        m = TITLE_LINE_RE.match(line)
        if m:
            title_feature = (m.group("feat") or "").strip()
            title_desc = m.group("desc").strip()
            break

    # 功能模块：优先从标题行的 <feature/feature> 提取
    feature_list = [f.strip() for f in title_feature.split("/") if f.strip()]
    if not feature_list:
        feature_list = [title_desc[:12]] if title_desc else []

    # 关键文本：用于判断具体改动语义
    key_text = " ".join(
        filter(
            None,
            [
                title_desc,
                fields.get("solution", ""),
                fields.get("change reason", ""),
                fields.get("root cause", ""),
                fields.get("rn description", ""),
                fields.get("test-proposal", ""),
            ],
        )
    )

    return {
        "feature_list": feature_list,
        "title_desc": title_desc,
        "fields": fields,
        "key_text": key_text,
        "has_test_proposal": bool(fields.get("test-proposal", "").lower().startswith("y")),
        "is_stress": fields.get("stress-test", "").lower().startswith("y"),
        "is_hw": fields.get("hw-test", "").lower().startswith("y"),
    }


def _build_case(summary: str, pre: str, steps: List[str], expects: List[str]) -> Dict[str, str]:
    return {
        "summary": summary,
        "pre_condition": pre,
        "test_step": "\n".join(f"{i}. {s}" for i, s in enumerate(steps, 1)),
        "expected_result": "\n".join(f"{i}. {s}" for i, s in enumerate(expects, 1)),
    }


def _describe_feature(feature: str, ctx: Dict[str, Any]) -> Dict[str, str]:
    """针对单个功能模块，结合上下文输出更具体的用例草稿。"""
    text = ctx["key_text"].lower()
    feat = feature.lower()
    title = ctx["title_desc"]

    # 运输模式相关
    if "运输" in feat or "transport" in feat:
        mode_name = f"{feature}模式" if not feature.endswith("模式") else feature
        return _build_case(
            summary=f"验证{mode_name}进入/退出及功耗影响",
            pre="模块正常开机并成功注网；功耗测试环境就绪，可测量模块电流。",
            steps=[
                f"触发模块进入{mode_name}（按项目定义的方式配置并启用）。",
                f"在{mode_name}下保持一段时间，观察模块状态及功耗。",
                f"触发退出{mode_name}，恢复常规工作状态。",
            ],
            expects=[
                f"模块成功进入{mode_name}，关键通信功能按设计进入低功耗或禁用状态。",
                f"{mode_name}期间电流符合设计预期，无异常发热或死机。",
                f"退出{mode_name}后模块能正常恢复注网并响应后续业务。",
            ],
        )

    # 日志功能，且提到 100k / 丢包 / 缓冲区 / post / 锁
    if "日志" in feat or "log" in feat:
        cases = []
        cases.append(
            _build_case(
                summary=f"验证{feature}大文件/大数据包上传完整性",
                pre="模块正常开机注网；日志服务器可用；准备约 100KB 大小的测试日志数据。",
                steps=[
                    "配置模块开启日志上传功能，并指向可用日志服务器。",
                    "触发模块产生并上传约 100KB 的日志数据包。",
                    "在服务器侧校验收到的日志内容与本地生成内容是否一致。",
                ],
                expects=[
                    "日志数据包完整上传到服务器，无丢包、无截断。",
                    "服务器侧文件大小与内容校验和与本地预期一致。",
                ],
            )
        )
        if any(k in text for k in ["1600", "缓冲区", "缓存", "post", "锁", "at_http"]):
            cases.append(
                _build_case(
                    summary=f"验证{feature}上传期间 AT 通道并发响应正常",
                    pre="模块正常开机注网；日志上传功能已启用；PC 端 AT 工具准备就绪。",
                    steps=[
                        "启动大日志数据上传，使模块处于 POST/写文件状态。",
                        "在日志上传过程中，持续向模块发送多条查询类 AT 指令。",
                        "观察 AT 响应时间、响应完整性和模块稳定性。",
                    ],
                    expects=[
                        "AT 指令在日志上传期间仍能正常响应，无长时间无响应或 ERROR。",
                        "日志上传完成后模块无死机、重启、异常锁死等现象。",
                        "上传数据完整，未因并发 AT 操作而丢包。",
                    ],
                )
            )
        if ctx["is_stress"]:
            cases.append(
                _build_case(
                    summary=f"压力测试：循环触发{feature}上传并叠加 AT 并发",
                    pre="模块正常开机注网；自动化脚本可循环触发日志上传并并发发送 AT。",
                    steps=[
                        f"循环触发{feature}上传 N 次（建议 ≥50 次），每次上传 50~100KB 数据。",
                        "每次上传期间并发执行多条 AT 查询指令。",
                        "持续运行并统计成功/失败次数及异常现象。",
                    ],
                    expects=[
                        "循环上传成功率 100%，无丢包、无截断。",
                        "AT 并发响应正常，无锁死、重启、内存泄漏。",
                    ],
                )
            )
        return cases[0] if len(cases) == 1 else cases

    # 网络/注网相关
    if any(k in feat for k in ["注网", "网络", "附着", "注册", "network", "attach", "registration"]):
        return _build_case(
            summary=f"验证{feature}在各种网络环境下的稳定性",
            pre="模块正常开机；SIM 卡已插入；测试环境可覆盖不同信号强度或不同运营商网络。",
            steps=[
                f"在标准信号环境下执行{feature}相关操作。",
                "将模块移动到弱信号/信号波动环境，重复上述操作。",
                "恢复正常信号，观察模块能否自动恢复。",
            ],
            expects=[
                f"标准环境下{feature}成功完成。",
                f"信号波动时{feature}不崩溃、不死机，恢复信号后能自动恢复。",
            ],
        )

    # AT 指令相关
    if "at" in feat or "指令" in feat or "命令" in feat:
        return _build_case(
            summary=f"验证{feature}指令响应符合规范",
            pre="模块正常开机注网；PC 端 AT 端口已连接。",
            steps=[
                f"按规范发送 {feature} 相关 AT 指令。",
                "检查返回内容及格式。",
                "边界参数下再次发送指令，观察容错行为。",
            ],
            expects=[
                "AT 指令返回格式正确，返回值与规范一致。",
                "非法/边界参数返回明确 ERROR 或指定提示，不导致模块异常。",
            ],
        )

    # 默认：基于 title_desc 生成通用但不过于空洞的用例
    short_desc = (title or ctx["fields"].get("solution", ""))[:30]
    return _build_case(
        summary=f"验证{feature}功能变更后的正确性",
        pre="模块正常开机，成功注网，相关前置配置已完成。",
        steps=[
            f"触发{feature}相关功能操作（{short_desc or '按变更描述执行'}）。",
            "观察模块输出、日志或 AT 返回结果。",
            "必要时执行相关回归检查，确认未引入副作用。",
        ],
        expects=[
            f"{feature}功能按预期工作，输出结果与变更描述一致。",
            "未发现异常重启、死机、内存泄漏或功能退化。",
        ],
    )


def ai_generate_test_cases(description: str) -> List[Dict[str, str]]:
    """【AI 环节】根据修改点描述智能生成测试用例草稿。

    改进点：
      - 解析 <Change Type>/<Solution>/<Change Reason>/<RN description>/<Test-Proposal> 等字段
      - 按标题行的功能模块拆分，针对每个模块生成有实质内容的用例
      - 识别关键字（运输、日志、100k、缓冲区、post、锁、AT 等）生成具体步骤和期望结果
      - 若标记了 Stress-Test，额外生成压力/并发用例
      - 若标记了 HW-Test，额外提示硬件测试项

    当前为本地规则化实现，无需联网；如需接入 LLM，可在此函数内替换生成逻辑。
    """
    ctx = _extract_context(description)
    if not ctx["feature_list"]:
        # 没有任何可识别的功能点时，基于整体描述生成一条通用用例
        body = strip_author_date(description)[:200]
        return [
            _build_case(
                summary="验证本次代码变更的功能正确性",
                pre="模块正常开机，成功注网，相关测试环境已就绪。",
                steps=[
                    "根据变更描述部署测试环境。",
                    f"执行与变更相关的功能验证：{body or '详见变更描述'}。",
                    "记录并比对实际结果与预期结果。",
                ],
                expects=[
                    "功能行为与变更描述一致。",
                    "无异常重启、死机、内存泄漏或功能退化。",
                ],
            )
        ]

    all_cases: List[Dict[str, str]] = []
    seen_summaries = set()
    for feature in ctx["feature_list"]:
        item_or_list = _describe_feature(feature, ctx)
        if isinstance(item_or_list, dict):
            items = [item_or_list]
        else:
            items = item_or_list
        for case in items:
            key = case["summary"]
            if key in seen_summaries:
                continue
            seen_summaries.add(key)
            all_cases.append(case)

    # HW-Test 标记：增加一条提示性用例（不替代硬件测试）
    if ctx["is_hw"]:
        all_cases.append(
            _build_case(
                summary="硬件相关检查：确认变更涉及的硬件接口/信号正常",
                pre="硬件测试环境就绪；示波器/电源等仪器已校准。",
                steps=[
                    "根据变更描述确认受影响的硬件接口或信号。",
                    "在典型工作条件下测量相关信号/功耗/时序。",
                    "与历史版本基线对比，确认无异常偏移。",
                ],
                expects=[
                    "硬件指标符合设计规格。",
                    "未发现异常纹波、功耗突增或时序违规。",
                ],
            )
        )

    return all_cases


def _create_case_with_retry(client: SmodClient, max_retry: int = 3, **kwargs) -> Any:
    """带重试地新建用例。

    仅对瞬时错误（HTTP 5xx / 网络异常 / 超时）重试；
    业务错误（如重复用例 code 30001）不重试，直接抛出，避免无效死循环。
    """
    last = None
    for attempt in range(1, max_retry + 1):
        try:
            return client.create_case(**kwargs)
        except Exception as e:
            last = e
            msg = str(e)
            # 业务级错误（重复用例等）无需重试
            if "30001" in msg or "duplicate" in msg.lower():
                raise
            if attempt < max_retry:
                print(f"    [重试] 新建用例失败（{msg[:80]}），第 {attempt} 次重试...")
                time.sleep(1.5 * attempt)
                continue
            raise
    raise last


def process_points_after_import(
    client: SmodClient,
    id_beta_ver: int,
    platform_code: str = "odmm",
    skip_no_req: bool = False,
) -> None:
    """导入 Excel 后，逐修改点处理测试用例与测试结果。

    规则：
      - 有需求：AI 生成测试用例（test_result=Test-in-Process），
                并把该修改点测试结果设为 Test-in-Process (ti)
      - 无需求（描述仅版本号）：不建用例，
                修改点测试结果设为 Blocked-NoRun (bnr)，
                备注设为「无需测试」
    """
    print(f"\n===== 步骤 7：逐修改点生成用例 & 设置测试结果 =====")
    result = client.list_points(id_beta_ver, page=1, size=500, platform_code=platform_code)
    records = result.get("records", []) if isinstance(result, dict) else []
    total = result.get("total", len(records)) if isinstance(result, dict) else len(records)
    print(f"该版本共 {total} 个修改点，开始处理...")

    n_with = n_without = 0
    for pt in records:
        pid = pt.get("id")
        desc = pt.get("description", "")

        if not has_requirement(desc):
            # 无需求：Blocked-NoRun + 备注
            client.set_point_test_result(pid, client.TR_BLOCKED_NORUN, platform_code)
            client.set_point_remark(pid, "无需测试", platform_code)
            n_without += 1
            print(f"  [无需求] 修改点 {pid} -> Blocked-NoRun + 备注『无需测试』")
            continue

        # 有需求：幂等——该修改点已有用例则跳过新建，仅确保测试结论正确
        existing = client.list_cases_for_point(pid, platform_code)
        if existing:
            print(f"  [有需求] 修改点 {pid} 已有 {len(existing)} 条用例，跳过新建（幂等）")
            client.set_point_test_result(pid, client.TR_IN_PROCESS, platform_code)
            n_with += 1
            continue

        # 有需求：AI 生成用例
        cases = ai_generate_test_cases(desc)
        for c in cases:
            code = client.get_new_case_code()
            _create_case_with_retry(
                client,
                code=code,
                summary=c["summary"],
                id_revise=pid,
                pre_condition=c.get("pre_condition", ""),
                test_step=c.get("test_step", ""),
                expected_result=c.get("expected_result", ""),
                test_result=client.TR_IN_PROCESS,
                platform_code=platform_code,
            )
        # 修改点自身测试结果 = Test-in-Process
        client.set_point_test_result(pid, client.TR_IN_PROCESS, platform_code)
        n_with += 1
        print(f"  [有需求] 修改点 {pid} -> 新建 {len(cases)} 条用例，测试状态 Test-in-Process")

    print(f"\n处理完成：有需求 {n_with} 个，无需求 {n_without} 个")


# =====================================================================
# 主工作流
# =====================================================================
def run_workflow(
    txt_path: str,
    version_code: str,
    access_token: Optional[str] = None,
    username: Optional[str] = None,
    password: Optional[str] = None,
    platform_code: str = "odmm",
    xlsx_path: Optional[str] = None,
    auto_create: bool = True,
) -> None:
    """完整工作流入口。

    登录方式二选一（优先用 token）：
      - access_token：浏览器复制注入
      - username + password：自动走 SSO 登录接口拿 token
    """

    # 1. txt -> xlsx
    print(f"\n===== 步骤 1：生成 Excel =====")
    xlsx_path = convert_txt_to_xlsx(txt_path, xlsx_path)

    # 2. 提取项目名
    project_name = parse_project_name(version_code)
    print(f"\n===== 步骤 2：解析版本号 =====")
    print(f"版本号：{version_code}")
    print(f"项目名：{project_name}")

    # 3. 登录 SMOD
    print(f"\n===== 步骤 3：登录 SMOD =====")
    client = SmodClient(platform_code=platform_code)
    try:
        if access_token:
            client.set_token(access_token)
        elif username and password:
            client.login_by_password(username, password)
        else:
            raise ValueError("必须提供 --token 或 --username + --password 之一")
        user = client.get_user()
        print(f"当前用户：{user.get('name')} ({user.get('email')})")
    except Exception as e:
        print(f"登录/鉴权失败：{e}")
        sys.exit(1)

    # 4. 推断平台
    print(f"\n===== 步骤 4：推断硬件平台 =====")
    id_plat_ver, name_plat_ver = infer_platform_from_history(client, project_name, platform_code)
    print(f"使用平台：{name_plat_ver} (id_plat_ver={id_plat_ver})")

    # 5. 检查版本是否已存在
    print(f"\n===== 步骤 5：检查版本是否存在 =====")
    result = client.search_projects(keyword=project_name, page=1, size=100, platform_code=platform_code)
    records = result.get("records", []) if isinstance(result, dict) else []
    matched = [r for r in records if str(r.get("code", "")).strip().lower() == version_code.strip().lower()]

    if matched:
        id_beta_ver = matched[0]["id"]
        print(f"版本已存在，id_beta_ver={id_beta_ver}，直接导入 Excel")
    else:
        if not auto_create:
            print(f"版本 {version_code} 不存在，且未开启自动创建，流程结束。")
            sys.exit(0)
        print(f"版本不存在，新建项目：{version_code}")
        client.create_project(code=version_code, id_plat_ver=id_plat_ver, platform_code=platform_code)
        # 再查一次获取 id_beta_ver
        result = client.search_projects(keyword=project_name, page=1, size=100, platform_code=platform_code)
        records = result.get("records", []) if isinstance(result, dict) else []
        matched = [r for r in records if str(r.get("code", "")).strip().lower() == version_code.strip().lower()]
        if not matched:
            raise SmodApiError("新建项目后仍查不到该版本，请检查接口响应。")
        id_beta_ver = matched[0]["id"]
        print(f"新建成功，id_beta_ver={id_beta_ver}")

    # 6. 导入 Excel
    print(f"\n===== 步骤 6：导入 Excel =====")
    data = client.import_points_excel(id_beta_ver=id_beta_ver, file_path=xlsx_path, platform_code=platform_code)
    print(f"导入成功：{data}")

    # 7. 逐修改点生成用例 & 设置测试结果
    process_points_after_import(
        client, id_beta_ver=id_beta_ver, platform_code=platform_code
    )


def main():
    ap = argparse.ArgumentParser(description="修改点 txt -> 生成 Excel -> 导入 SMOD")
    ap.add_argument("txt", nargs="?", default=None, help="修改点 txt 路径")
    ap.add_argument("version", nargs="?", default=None, help="版本号，如 QDM559_STM32G0B0_APP_01.001.01.001_V19")
    ap.add_argument("--token", default=None, help="浏览器复制的 access_token（去掉 bearer 前缀）")
    ap.add_argument("--username", default=None, help="SMOD/SSO 用户名，如 Venus.Li@ikotek.com")
    ap.add_argument("--password", default=None, help="SMOD/SSO 密码")
    ap.add_argument("--platform", default="odmm", help="平台代码，默认 odmm")
    ap.add_argument("--xlsx", default=None, help="生成的 xlsx 路径，默认与 txt 同名")
    ap.add_argument("--no-create", action="store_true", help="若版本不存在则不新建，直接退出")
    ap.add_argument("--delete-cases", default=None,
                    help="仅删除用例模式：逗号分隔的用例 id 列表，登录后直接删除并退出（如 1851245,1821972）")
    args = ap.parse_args()

    # ---- 仅删除用例模式（不依赖 txt/版本号，跳过交互补全）----
    if args.delete_cases:
        ids = [int(x.strip()) for x in args.delete_cases.split(",") if x.strip()]
        client = SmodClient(platform_code=args.platform)
        if args.token:
            client.set_token(args.token)
        elif args.username and args.password:
            client.login_by_password(args.username, args.password)
        else:
            import getpass
            u = input("SMOD/SSO 用户名: ").strip() or "Venus.Li@ikotek.com"
            p = getpass.getpass("SMOD/SSO 密码: ")
            client.login_by_password(u, p)
        print(f"删除用例 id 列表：{ids}")
        print("删除结果：", client.delete_cases(ids, platform_code=args.platform))
        sys.exit(0)

    # 缺必填参数时弹窗/交互式补全
    args = prompt_missing_args(args)

    # 弹窗后仍缺失则退出
    if not args.txt:
        ap.error("必须提供修改点 txt 路径")
    if not args.version:
        ap.error("必须提供版本号")
    if not args.token and not (args.username and args.password):
        ap.error("必须提供 --token 或同时提供 --username + --password")

    run_workflow(
        txt_path=args.txt,
        version_code=args.version,
        access_token=args.token,
        username=args.username,
        password=args.password,
        platform_code=args.platform,
        xlsx_path=args.xlsx,
        auto_create=not args.no_create,
    )


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
changelist_to_xlsx.py
=====================
将 git changelist 文本（形如 `git log` / 自定义修改点导出格式）转换成
"spliter_tool" 风格的 Excel 修改点清单。

输入 txt 每个 commit 块格式（与 spliter_tool_v3.4 输出一致）：

    commit <hash>
    Author: <name> <email>
    Date:   <date>

        <提交信息，可含 <Field>: value 字段>

转换规则（已与原工具输出逐字节核对）：
  - 列 A 修改点编号   = commit 的 hash
  - 列 B 修改点描述   = 去掉首行 "commit <hash>" 后的整段内容（Author/Date/空行/缩进正文），去掉末尾换行
  - 列 C 修改点分类   = 由正文中的 <Change Type>: 字段推导：
        new requirements        -> Newly-added Customer Requirements
        bug fix / inner bug fix -> Internal Defect Requirements
        （缺失或嵌在 merge 单行中无法解析 -> 留空）
  - 列 D~J（执行人/测试结果/评审状态/备注/Test-Proposal/Stress-Test/HW-Test）
                           = 留空，供后续人工填写（与原工具行为一致）

用法：
    python changelist_to_xlsx.py input.txt [output.xlsx]

若不指定 output.xlsx，则在与输入同目录生成 <输入名>.xlsx（与 txt 同名）。
"""

import argparse
import os
import re
import sys

# 自动安装依赖（仅固定包 openpyxl），保证脚本自包含
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

from collections import Counter

# ----------------------------- 配置 -----------------------------
HEADERS = [
    "修改点编号", "修改点描述", "修改点分类", "执行人", "测试结果",
    "评审状态", "备注", "Test-Proposal", "Stress-Test", "HW-Test",
]

# <Change Type> 取值 -> Excel 中的"修改点分类"
CATEGORY_MAP = {
    "new requirements": "Newly-added Customer Requirements",
    "bug fix": "Internal Defect Requirements",
    "inner bug fix": "Internal Defect Requirements",
}

# 表头样式（与原工具一致：加粗 + 灰色填充 00969696）
HEADER_FILL = PatternFill(fill_type="solid", fgColor="969696")
HEADER_FONT = Font(bold=True)
DESC_ALIGN = Alignment(wrap_text=True, vertical="top")


# --------------------------- 解析逻辑 ---------------------------
def read_text(path: str) -> str:
    """读取文本，优先 utf-8，失败回退 gbk/utf-8-sig（兼容中文 Windows git 导出）。"""
    for enc in ("utf-8", "gbk", "utf-8-sig"):
        try:
            with open(path, "r", encoding=enc) as f:
                return f.read()
        except (UnicodeDecodeError, LookupError):
            continue
    with open(path, "rb") as f:
        return f.read().decode("utf-8", errors="ignore")


def derive_category(block: str):
    """从 commit 正文推导修改点分类。

    与原工具一致：仅在行首（允许前导空白）出现 `<Change Type>:` 时才解析，
    因此 merge 提交里嵌在单行中间的 <Change Type> 不会被识别（返回 None）。
    """
    for line in block.splitlines():
        m = re.match(r"^\s*<([^>]+)>\s*:\s*(.*)$", line)
        if m and m.group(1).strip().lower() == "change type":
            val = m.group(2).strip().lower()
            for key, label in CATEGORY_MAP.items():
                if val == key or val == key.replace(" ", ""):
                    return label
            if "new requirements" in val:
                return CATEGORY_MAP["new requirements"]
            if "bug fix" in val:
                return CATEGORY_MAP["bug fix"]
    return None


def parse_commits(text: str):
    """把整段文本按 '^commit ' 切成多个 commit。

    返回 [(hash, description_block, category), ...]
    description_block = 去掉首行 "commit <hash>" 后的内容，并去除末尾换行。
    """
    commits = []
    for chunk in re.split(r"(?m)^commit ", text)[1:]:
        lines = chunk.splitlines()
        if not lines:
            continue
        commit_hash = lines[0].strip()
        block = "\n".join(lines[1:]).rstrip("\r\n")
        commits.append((commit_hash, block, derive_category(block)))
    return commits


# --------------------------- 生成 Excel ---------------------------
def build_workbook(commits):
    wb = Workbook()
    ws = wb.active
    ws.title = "修改点"

    for col, name in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center")

    for ridx, (commit_hash, block, category) in enumerate(commits, start=2):
        ws.cell(row=ridx, column=1, value=commit_hash)
        desc_cell = ws.cell(row=ridx, column=2, value=block)
        desc_cell.alignment = DESC_ALIGN
        ws.cell(row=ridx, column=3, value=category)
        # 列 D~J 留空

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 32
    for col in ("D", "E", "F", "G", "H", "I", "J"):
        ws.column_dimensions[col].width = 14

    return wb


def main():
    ap = argparse.ArgumentParser(description="changelist txt -> 修改点 Excel")
    ap.add_argument("input", help="输入 changelist .txt 路径")
    ap.add_argument("output", nargs="?", default=None, help="输出 .xlsx 路径（默认与输入同目录同名）")
    args = ap.parse_args()

    if not os.path.isfile(args.input):
        sys.exit(f"[错误] 找不到输入文件: {args.input}")

    out_path = args.output or os.path.splitext(args.input)[0] + ".xlsx"

    text = read_text(args.input)
    commits = parse_commits(text)
    if not commits:
        sys.exit("[警告] 未在输入中解析到任何 commit（确认是否为 git changelist 格式）。")

    wb = build_workbook(commits)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)

    cat_counter = Counter(c[2] or "(空)" for c in commits)
    print(f"[完成] 已解析 {len(commits)} 个 commit")
    for k, v in cat_counter.items():
        print(f"        {k}: {v}")
    print(f"[输出] {out_path}")


if __name__ == "__main__":
    main()

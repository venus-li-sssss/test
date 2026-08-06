#!/usr/bin/env python3
"""
Robust Excel parser for CAN protocol documents.
Handles all edge cases found in real-world CAN protocol Excel files.

Edge cases handled:
1. start_byte=None (signals share same byte as previous signal)
2. Bitfield signals (multiple 1-bit signals in same byte)
3. Merged cells in Excel
4. Incomplete signal definitions
5. Multiple pages for same CAN ID
"""

import openpyxl
import json
import re
from typing import Dict, List, Any, Optional


def safe_float(val: Any, default: float = 0.0) -> float:
    """Safely convert value to float."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.strip()
        if val in ['-', '', 'N/A', 'None', 'none']:
            return default
        try:
            return float(val)
        except ValueError:
            return default
    return default


def parse_bit_position(bit_pos_str: Any) -> tuple[Optional[int], Optional[int]]:
    """
    Parse bit position string like 'Bit0-15', 'Bit0', 'Bit32-63' etc.
    Returns (start_bit_offset, bit_length).
    
    Examples:
    - 'Bit0-15' -> (0, 16)
    - 'Bit0' -> (0, 1)
    - 'Bit16-31' -> (16, 16)
    """
    if bit_pos_str is None:
        return None, None
    
    bit_pos_str = str(bit_pos_str).strip()
    
    # Handle range like "Bit0-15"
    if '-' in bit_pos_str:
        match = re.search(r'Bit(\d+)-(\d+)', bit_pos_str)
        if match:
            start = int(match.group(1))
            end = int(match.group(2))
            return start, (end - start + 1)
    
    # Handle single bit like "Bit0"
    match = re.search(r'Bit(\d+)', bit_pos_str)
    if match:
        return int(match.group(1)), 1
    
    return None, None


def parse_value_descriptions(desc_str: Any) -> Dict[str, str]:
    """
    Parse signal value description string into a mapping dict.

    Handles formats found in protocol Excel files:
    - "0x00:普通日期\\n0x01:春节\\n0x02:元宵节"
    - "00:普通日期, 01:春节, 02:元宵节"
    - "0:工作日, 1:休息日"

    Returns:
        Dict mapping raw value string (as key) to human-readable meaning.
        Keys are normalized to lowercase hex strings without '0x' prefix.
    """
    if desc_str is None:
        return {}

    result = {}
    text = str(desc_str).strip()

    # Split by newlines, commas, or semicolons
    entries = re.split(r'[\n,;]+', text)

    for entry in entries:
        entry = entry.strip()
        if not entry or ':' not in entry:
            continue

        parts = entry.split(':', 1)
        if len(parts) != 2:
            continue

        raw_key = parts[0].strip()
        meaning = parts[1].strip()

        if not raw_key or not meaning:
            continue

        # Normalize key: remove 0x/0X prefix, lowercase
        norm_key = raw_key.lower().replace('0x', '').replace('0x', '')
        result[norm_key] = meaning

    return result


def parse_excel_can_protocol(excel_path: str) -> Dict[int, Dict[str, Any]]:
    """
    Parse Excel CAN protocol document and extract all message definitions.
    
    Args:
        excel_path: Path to Excel file
    
    Returns:
        Dict mapping CAN ID (int) to message config dict
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    all_messages = {}
    
    for sheet_name in wb.sheetnames:
        # Skip example sheets
        if '示例' in sheet_name or 'example' in sheet_name.lower():
            continue
        
        ws = wb[sheet_name]
        
        # Find header row by looking for specific column headers
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Check if this row contains the expected headers
            if row[0] == 'Num\n序号' or (isinstance(row[0], (int, float)) and row[0] == 1):
                # Verify this is actually the header row by checking other columns
                if row[7] == 'Signal Name\n信号名称' or 'Signal' in str(row[7]):
                    header_row = i
                    break
        
        if header_row is None:
            print(f"Warning: Could not find header row in sheet {sheet_name}")
            continue
        
        # Parse messages
        current_msg = None
        current_byte = None  # Track current byte for start_byte=None signals
        
        for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
            # Check if this is a new message (column A has a number)
            if row[0] is not None and isinstance(row[0], (int, float)):
                # Save previous message
                if current_msg is not None:
                    _validate_message(current_msg)
                    all_messages[current_msg['can_id']] = current_msg
                
                # Start new message
                can_id = _parse_can_id(row[4])
                if can_id is not None:
                    current_msg = {
                        'can_id': can_id,
                        'msg_name': row[1],
                        'direction': row[2],
                        'cycle_time': row[5],
                        'dlc': _parse_dlc(row[6]),
                        'signals': [],
                    }
                    current_byte = None
            
            # Add signal
            if current_msg is not None and row[7] is not None:
                signal_name = row[7]
                signal_desc = row[8] if len(row) > 8 and row[8] else ''   # 信号描述 (列I)
                start_byte = row[9]
                bit_position = row[10]
                bit_length = row[11]

                # Signal value description column (信号值描述) - typically column N
                # Try to find it dynamically: look for a column containing "0x00:" patterns
                value_desc_col = None
                for col_idx in range(13, min(len(row), 20)):
                    cell_val = row[col_idx]
                    if cell_val and isinstance(cell_val, str) and re.search(r'0x[0-9a-fA-F]+\s*:', cell_val):
                        value_desc_col = col_idx
                        break
                value_description = row[value_desc_col] if value_desc_col is not None and value_desc_col < len(row) else None
                
                # Debug output
                # print(f"Processing signal: {signal_name}, start_byte={start_byte}, bit_position={bit_position}")
                
                # Calculate start_bit
                if start_byte is not None:
                    try:
                        current_byte = int(start_byte)
                    except (ValueError, TypeError):
                        pass
                
                # Parse bit position to get bit offset
                bit_offset, parsed_length = parse_bit_position(bit_position)
                
                # Use parsed length if bit_length is not specified
                if bit_length is not None:
                    try:
                        bit_length_int = int(bit_length)
                    except (ValueError, TypeError):
                        bit_length_int = parsed_length
                else:
                    bit_length_int = parsed_length
                
                # Calculate start_bit
                if current_byte is not None and bit_offset is not None:
                    start_bit = current_byte * 8 + bit_offset
                else:
                    start_bit = None
                
                if start_bit is not None and bit_length_int is not None:
                    # Determine byte order
                    byte_order_str = row[12]
                    byte_order = 'intel'
                    if byte_order_str and 'motorola' in str(byte_order_str).lower():
                        byte_order = 'motorola'
                    
                    # Determine value type
                    value_type = 'unsigned'
                    min_val = row[19]
                    if min_val is not None:
                        try:
                            if safe_float(min_val) < 0:
                                value_type = 'signed'
                        except (ValueError, TypeError):
                            pass
                    
                    signal = {
                        'name': _clean_signal_name(signal_name),
                        'description': str(signal_desc).strip() if signal_desc else '',
                        'start_bit': start_bit,
                        'bit_length': bit_length_int,
                        'byte_order': byte_order,
                        'value_type': value_type,
                        'factor': safe_float(row[17], 1.0),
                        'offset': safe_float(row[18], 0.0),
                        'unit': row[21] if len(row) > 21 and row[21] else '',
                        'value_descriptions': parse_value_descriptions(value_description),
                    }
                    
                    current_msg['signals'].append(signal)
        
        # Don't forget last message
        if current_msg is not None:
            _validate_message(current_msg)
            all_messages[current_msg['can_id']] = current_msg
    
    return all_messages


def _parse_can_id(id_value: Any) -> Optional[int]:
    """Parse CAN ID from various formats."""
    if id_value is None:
        return None
    
    id_str = str(id_value).strip()
    
    # Remove any non-hex characters except 'x' and 'X'
    id_str = re.sub(r'[^0-9a-fA-FxX]', '', id_str)
    
    try:
        if id_str.lower().startswith('0x'):
            return int(id_str, 16)
        else:
            return int(id_str)
    except ValueError:
        return None


def _parse_dlc(dlc_value: Any) -> int:
    """Parse DLC value."""
    if dlc_value is None:
        return 8
    
    try:
        return int(dlc_value)
    except (ValueError, TypeError):
        return 8


def _clean_signal_name(name: str) -> str:
    """Clean signal name for use as Python variable."""
    name = str(name).strip()
    name = re.sub(r'[^\w\s]', '_', name)
    name = re.sub(r'\s+', '_', name)
    return name.lower()


def _validate_message(msg: Dict[str, Any]) -> None:
    """
    Validate and fix signal definitions in a message.
    
    Fixes:
    1. Signals that exceed DLC
    2. Duplicate signal names
    """
    dlc = msg['dlc']
    max_bit = dlc * 8
    
    # Remove signals that exceed DLC
    valid_signals = []
    for sig in msg['signals']:
        if sig['start_bit'] + sig['bit_length'] <= max_bit:
            valid_signals.append(sig)
        else:
            print(f"Warning: Signal {sig['name']} in {msg['msg_name']} "
                  f"exceeds DLC (start_bit={sig['start_bit']}, "
                  f"length={sig['bit_length']}, DLC={dlc})")
    
    # Fix duplicate signal names
    name_count = {}
    for sig in valid_signals:
        name = sig['name']
        if name in name_count:
            name_count[name] += 1
            sig['name'] = f"{name}_{name_count[name]}"
        else:
            name_count[name] = 1
    
    msg['signals'] = valid_signals


def export_to_json(messages: Dict[int, Dict], output_path: str) -> None:
    """Export parsed messages to JSON config file."""
    config = {}
    
    for can_id, msg in messages.items():
        config[can_id] = {
            'msg_name': msg['msg_name'],
            'direction': msg.get('direction', ''),
            'cycle_time': msg.get('cycle_time'),
            'dlc': msg['dlc'],
            'signals': [
                {
                    'name': sig['name'],
                    'description': sig.get('description', ''),
                    'start_bit': sig['start_bit'],
                    'bit_length': sig['bit_length'],
                    'byte_order': sig['byte_order'],
                    'value_type': sig['value_type'],
                    'factor': sig['factor'],
                    'offset': sig['offset'],
                    'unit': sig['unit'],
                    'value_descriptions': sig.get('value_descriptions', {}),
                }
                for sig in msg['signals']
            ]
        }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)
    
    print(f"Exported {len(config)} messages to {output_path}")


def main():
    """Main function for testing."""
    import sys
    
    if len(sys.argv) < 3:
        print("Usage: python excel_parser.py <excel_path> <output_json_path>")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2]
    
    print(f"Parsing {excel_path}...")
    messages = parse_excel_can_protocol(excel_path)
    
    print(f"Parsed {len(messages)} messages")
    
    # Print summary
    for can_id, msg in list(messages.items())[:5]:
        print(f"\n0x{can_id:08X}: {msg['msg_name']}")
        print(f"  DLC: {msg['dlc']}, Signals: {len(msg['signals'])}")
        for sig in msg['signals'][:3]:
            print(f"    - {sig['name']}: start_bit={sig['start_bit']}, "
                  f"length={sig['bit_length']}")
    
    export_to_json(messages, output_path)


if __name__ == '__main__':
    main()

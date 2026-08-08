# -*- coding: utf-8 -*-
"""
清空全表所有“未执行却被填了结果”的用例数据。
判定标准：实测结果列非空，但实测记录列和 Blocked-NoRun 原因列均为空。
清除内容：实测结果、Blocked 原因、实测记录、备注；同时移除该行图片、恢复默认样式。
"""
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

F = r"C:\Users\venus.li\WorkBuddy\2026-08-07-16-59-49\IK-QP-012-07_ODM_QDM522_TestCase_整车_V4_4G_Only_executed.xlsx"

wb = openpyxl.load_workbook(F, data_only=False)

blank_fill = PatternFill()
default_font = Font()
default_align = Alignment()

for ws in wb.worksheets:
    if ws.max_row < 13 or ws.max_column < 10:
        continue

    # 定位标题行
    header_row = None
    for r in range(1, min(15, ws.max_row + 1)):
        for c in range(1, min(ws.max_column + 1, 15)):
            v = ws.cell(r, c).value
            if v and "实测结果" in str(v):
                header_row = r
                break
        if header_row:
            break
    if not header_row:
        continue

    cols = {}
    for c in range(1, min(ws.max_column + 1, 22)):
        v = str(ws.cell(header_row, c).value or "")
        if "用例编号" in v or "Case ID" in v:
            cols["case"] = c
        if "实测结果" in v:
            cols["result"] = c
        if "Blocked" in v or "NoRun" in v:
            cols["blocked"] = c
        if "实测记录" in v:
            cols["record"] = c
        if "备注" in v:
            cols["note"] = c

    if not cols.get("case") or not cols.get("result"):
        continue

    case_col = cols["case"]
    result_col = cols["result"]
    blocked_col = cols.get("blocked")
    record_col = cols.get("record")
    note_col = cols.get("note")

    purged = 0
    for r in range(header_row + 1, ws.max_row + 1):
        case = ws.cell(r, case_col).value
        if not case or "说明" in str(case) or "测试说明" in str(case):
            continue

        result = ws.cell(r, result_col).value
        blocked = ws.cell(r, blocked_col).value if blocked_col else None
        record = ws.cell(r, record_col).value if record_col else None

        # 只要实测结果列有内容，但无实测记录且无 Blocked 原因 → 视为未执行却填了结果
        if result and str(result).strip() and not record and not blocked:
            # 清空结果、原因、记录、备注
            for c in (result_col, blocked_col, record_col, note_col):
                if not c:
                    continue
                cell = ws.cell(r, c)
                cell.value = None
                cell.fill = blank_fill
                cell.font = default_font
                cell.alignment = default_align

            # 删除该行内所有图片（openpyxl 的 OneCellAnchor / TwoCellAnchor 均按锚点顶行判断）
            remaining = []
            for img in getattr(ws, "_images", []):
                anchor = getattr(img, "anchor", None)
                if anchor is None:
                    remaining.append(img)
                    continue
                # 获取锚点起始行（_from 属性存在多种命名）
                top_row = None
                if hasattr(anchor, "_from"):
                    frm = anchor._from
                    top_row = getattr(frm, "row", None) or getattr(frm, "rowOff", None)
                    if top_row is None and hasattr(frm, "_row"):
                        top_row = frm._row
                elif hasattr(anchor, "row"):
                    top_row = anchor.row
                if top_row == r - 1:  # openpyxl 内部行号 0-based
                    continue  # 删除属于该行的图片
                remaining.append(img)
            ws._images = remaining

            purged += 1

    if purged:
        print(f"{ws.title}: 清空 {purged} 行未执行数据")

wb.save(F)
print(f"saved {F}")

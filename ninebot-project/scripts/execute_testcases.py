#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
execute_testcases.py — 测试用例执行工作流引擎（ninebot-project skill 配套）

工作流（与 SKILL.md §8.0 对齐）：
  1. next    : 读取下一条【未执行】的用例组（按「用例编号」分组，组内多行视为同一条）
  2. (Agent 按用例步骤，用 device_control.py 的 go_to_page / cmd / precise_range_status 等执行，截图取证)
  3. record  : 回填该用例组的结果/记录/备注，并把 3 张取证截图嵌入 P/Q/R；同编号多行一起填
  4. 回到 1，直到 status 显示 done==total

列约定（自动探测表头，找不到则新增 P/Q/R 表头）：
  J=实测结果(verdict)  K=Blocked-NoRun原因  L=实测记录  M=备注
  P/Q/R = 3 张取证截图（整图等比缩放适配行，行高 = 图高*0.75pt，不裁剪内容）

用法：
  python execute_testcases.py next   --xlsx X.xlsx [--sheet 精准续航] [--state st.json]
  python execute_testcases.py record --xlsx X.xlsx --sheet 精准续航 --id 精准续航-01-002 \
        --verdict PASS --record "..." --remark "..." \
        --imgs before.png after.png result.png
  python execute_testcases.py status --xlsx X.xlsx [--sheet 精准续航]
  python execute_testcases.py reset  --xlsx X.xlsx --sheet 精准续航 [--id 精准续航-01-002]  # 清空重测
"""
import argparse, json, os, re, sys, tempfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

ID_RE = re.compile(r'^[^/\s]+-\d+-\d+$')

def find_header(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v and '用例编号' in str(v):
                return r
    return None

def col_letter(idx):
    from openpyxl.utils import get_column_letter
    return get_column_letter(idx)

def ensure_pqr(ws, hr):
    """确保 P/Q/R (16/17/18) 有截图表头；没有就补。返回 {col_idx:letter}。"""
    need = {16: '截图1', 17: '截图2', 18: '截图3'}
    for c, name in need.items():
        if ws.cell(hr, c).value in (None, ''):
            ws.cell(hr, c, name)
    return need

def read_groups(ws, hr):
    """返回 [(用例编号, [row_indices])]，去掉空行/非用例行。"""
    groups = {}
    for r in range(hr + 1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if b is None:
            continue
        bs = str(b).strip()
        if ID_RE.match(bs):
            groups.setdefault(bs, []).append(r)
    return groups

def is_done(ws, hr, rows):
    """J 列(10)有内容即视为已执行。"""
    for r in rows:
        if ws.cell(r, 10).value not in (None, ''):
            return True
    return False

def state_path(xlsx, st):
    return st or (xlsx + '.tcstate.json')

def load_state(xlsx, st):
    p = state_path(xlsx, st)
    if os.path.exists(p):
        try:
            return json.load(open(p, encoding='utf-8'))
        except Exception:
            return {}
    return {}

def save_state(xlsx, st, data):
    p = state_path(xlsx, st)
    json.dump(data, open(p, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)

def cmd_next(args):
    wb = load_workbook(args.xlsx)
    sheet = args.sheet or wb.sheetnames[0]
    ws = wb[sheet]
    hr = find_header(ws) or 10
    groups = read_groups(ws, hr)
    state = load_state(args.xlsx, args.state).get(sheet, [])
    done_ids = set(state)
    for cid, rows in groups.items():
        if cid in done_ids:
            continue
        if is_done(ws, hr, rows):
            # 已填 J 但未登记到 state（如模板预填）→ 也跳过，避免重复
            continue
        # 打印该组第一条为主，并提示同组其他行
        r0 = rows[0]
        def g(c): 
            v = ws.cell(r0, c).value; return '' if v is None else str(v)
        print(f"=== NEXT 用例组: {cid} （同编号行: {rows}）===")
        print(f"模块   : {g(3)}")
        print(f"测试项 : {g(4)}")
        print(f"优先级 : {g(5)}")
        print(f"测试内容: {g(6)}")
        print(f"前置条件: {g(7)}")
        print(f"执行步骤: {g(8)}")
        print(f"期望结果: {g(9)}")
        print("--- 请按步骤用 device_control.py 执行，截图取证，再用 record 回填 ---")
        return
    print("ALL DONE: 没有未执行的用例组了。")

_EMBED_MAX_W = 240          # 嵌入图最大宽度（保持纵横比）
_EMBED_MAX_H = 530           # 嵌入图最大高度（对应 397pt 行高，适配 Excel 409pt 上限）
_EMBED_PAD   = 14            # 行高 = 图高(pt) + padding

def _resize_for_embed(im):
    """整图等比缩放塞进 _EMBED_MAX_W x _EMBED_MAX_H 框（不裁剪）。"""
    w, h = im.size
    s = min(_EMBED_MAX_W / w, _EMBED_MAX_H / h)
    if s >= 1:
        return w, h
    return int(round(w * s)), int(round(h * s))

def embed(ws, cell_ref, imgpath, row):
    """嵌入取证图：整图【等比缩放】到适配一行的尺寸（不裁剪），并按图高设置行高，
    保证图片刚好落在所在行内、不溢出到下一行。
    行高(pt) = 图高(px) * 0.75  +  padding（96dpi: 1px=0.75pt）。"""
    im = PILImage.open(imgpath).convert('RGB')
    nw, nh = _resize_for_embed(im)
    if (nw, nh) != im.size:
        im = im.resize((nw, nh), PILImage.LANCZOS)
    fd, tmp = tempfile.mkstemp(suffix='.png')
    os.close(fd)
    im.save(tmp, optimize=True)
    img = XLImage(tmp)
    img.width, img.height = nw, nh          # 显式设定嵌入像素尺寸
    img.anchor = cell_ref
    ws.add_image(img)
    # 行高 = 图高(pt) + padding，且有图行取最大值，避免文字/图片重叠
    target_pt = nh * 0.75 + _EMBED_PAD
    cur = ws.row_dimensions[row].height or 0
    if cur < target_pt:
        ws.row_dimensions[row].height = round(target_pt, 1)

def cmd_record(args):
    wb = load_workbook(args.xlsx)
    sheet = args.sheet or wb.sheetnames[0]
    ws = wb[sheet]
    hr = find_header(ws) or 10
    ensure_pqr(ws, hr)
    groups = read_groups(ws, hr)
    if args.id not in groups:
        print(f"[ERROR] 用例编号 {args.id} 不在 sheet「{sheet}」中。可选: {list(groups.keys())}")
        sys.exit(1)
    rows = groups[args.id]
    imgs = args.imgs or []
    for i, r in enumerate(rows):
        ws.cell(r, 10, args.verdict)                 # J 实测结果
        if args.verdict and args.verdict.strip().lower().startswith('block'):
            ws.cell(r, 11).value = args.blockcause or ''    # K
        else:
            ws.cell(r, 11).value = None
        ws.cell(r, 12, args.record or '')            # L 实测记录
        ws.cell(r, 13, args.remark or '')            # M 备注
        # 截图嵌入第一行（同组多行共用同一组图，避免重复）
        if i == 0:
            if args.col:
                # 单图模式：把 imgs[0] 贴到指定列（如 19=S 放平台指令日志截图）
                if imgs and os.path.exists(imgs[0]):
                    embed(ws, f"{col_letter(args.col)}{r}", imgs[0], r)
            else:
                for j, imgp in enumerate(imgs[:4]):
                    if imgp and os.path.exists(imgp):
                        embed(ws, f"{col_letter(16+j)}{r}", imgp, r)
    wb.save(args.xlsx)
    # 登记 state
    state = load_state(args.xlsx, args.state)
    done = state.setdefault(sheet, [])
    if args.id not in done:
        done.append(args.id)
    save_state(args.xlsx, args.state, state)
    print(f"[OK] 已回填 {args.id} -> 行 {rows}，verdict={args.verdict}，截图 {len(imgs[:3])} 张嵌入 P/Q/R。")

def cmd_status(args):
    wb = load_workbook(args.xlsx)
    sheet = args.sheet or wb.sheetnames[0]
    ws = wb[sheet]
    hr = find_header(ws) or 10
    groups = read_groups(ws, hr)
    state = load_state(args.xlsx, args.state).get(sheet, [])
    total = len(groups)
    done = sum(1 for cid in groups if cid in state or is_done(ws, hr, groups[cid]))
    print(f"sheet「{sheet}」: 已执行 {done}/{total}")
    for cid, rows in groups.items():
        mark = '✓' if (cid in state or is_done(ws, hr, rows)) else '·'
        print(f"  {mark} {cid}  (行 {rows})")

def cmd_reset(args):
    wb = load_workbook(args.xlsx)
    sheet = args.sheet or wb.sheetnames[0]
    ws = wb[sheet]
    hr = find_header(ws) or 10
    groups = read_groups(ws, hr)
    targets = [args.id] if args.id else list(groups.keys())
    for cid in targets:
        if cid not in groups:
            continue
        for r in groups[cid]:
            ws.cell(r, 10).value = None; ws.cell(r, 11).value = None
            ws.cell(r, 12).value = None; ws.cell(r, 13).value = None
            # 清除 P/Q/R/S(16/17/18/19) 列图片，兼容 OneCellAnchor(无 ref)和 TwoCellAnchor(有 ref)
            for c in (16, 17, 18, 19):
                for img in list(ws._images):
                    a = img.anchor
                    hit = False
                    if hasattr(a, 'ref') and a.ref and a.ref.startswith(col_letter(c) + str(r)):
                        hit = True
                    elif hasattr(a, '_from') and a._from and a._from.col == c - 1 and a._from.row == r - 1:
                        hit = True
                    if hit:
                        ws._images.remove(img)
    wb.save(args.xlsx)
    state = load_state(args.xlsx, args.state)
    if sheet in state:
        if args.id:
            state[sheet] = [x for x in state[sheet] if x != args.id]
        else:
            state[sheet] = []
        save_state(args.xlsx, args.state, state)
    print(f"[OK] 已清空 {targets} 的 J/K/L/M/P/Q/R，可重测。")

def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest='cmd', required=True)
    for name in ('next', 'record', 'status', 'reset'):
        sp = sub.add_parser(name)
        sp.add_argument('--xlsx', required=True)
        sp.add_argument('--sheet', default=None)
        sp.add_argument('--state', default=None)
        if name == 'record':
            sp.add_argument('--id', required=True)
            sp.add_argument('--verdict', required=True)
            sp.add_argument('--record', default='')
            sp.add_argument('--remark', default='')
            sp.add_argument('--blockcause', default='')
            sp.add_argument('--imgs', nargs='*', default=[])
            sp.add_argument('--col', type=int, default=None,
                           help='单图模式: 把 imgs[0] 贴到指定列(如 19=S 放平台指令日志截图)')
            sp.add_argument('--only-embed', action='store_true',
                           help='只贴图, 不覆盖 J/K/L/M 文字列')
        if name == 'reset':
            sp.add_argument('--id', default=None)
    args = ap.parse_args()
    {'next': cmd_next, 'record': cmd_record, 'status': cmd_status, 'reset': cmd_reset}[args.cmd](args)

if __name__ == '__main__':
    main()
